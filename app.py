"""
PO PDF Extractor
-----------------
Upload a Purchase Order PDF (text-based OR scanned), let Claude (AI) read it,
preview/edit the extracted fields, then download as Excel.

Works for both:
  - True text PDFs (extracts text directly, sends text to Claude)
  - Scanned/image PDFs (renders page as image, sends image to Claude vision)
"""

import io
import json
import base64

import streamlit as st
import pandas as pd
import pdfplumber
from pdf2image import convert_from_bytes
from openpyxl import Workbook
from openpyxl.styles import Font
import anthropic


# ---------- CONFIG ----------

FIELDS = [
    "record_no",
    "item_code",
    "part_name",
    "part_no",
    "supplier",
    "total_amount",
    "requester",
    "po_date",
]

FIELD_LABELS = {
    "record_no": "Record No",
    "item_code": "Item Code",
    "part_name": "Part Name",
    "part_no": "Part No",
    "supplier": "Supplier",
    "total_amount": "Total Amount",
    "requester": "Requester",
    "po_date": "PO Date",
}

MODEL = "claude-sonnet-4-6"

EXTRACTION_PROMPT = (
    "You are reading a Purchase Order (PO) document. "
    "Extract the following fields for EVERY line item on the PO.\n\n"
    "Fields to extract for each line item row:\n"
    "- record_no: the line item number (an integer such as one, two, three, and so on)\n"
    "- item_code: the item or material code for that line\n"
    "- part_name: the description or name of the part on that line\n"
    "- part_no: the secondary part number on that line, if shown. Leave blank if not present.\n"
    "- supplier: the vendor or supplier company name (this is usually the SAME for every "
    "line item on one PO - it is the company name in the letterhead or address block, "
    "NOT the buyer's own company)\n"
    "- total_amount: the total amount for that line item (the final number in the amount "
    "column, usually the rightmost column)\n"
    "- requester: the name of the person who requested or raised the PO (often near the "
    "PO number or date block, sometimes labeled or just a first name)\n"
    "- po_date: the PO date (the order date, NOT a delivery date - usually near the PO number)\n\n"
    "Important rules:\n"
    "- Return ONE entry per line item row in the document.\n"
    "- po_date, supplier, and requester are typically the same across all line items in a "
    "single PO - repeat them on every row.\n"
    "- If a field cannot be found for a given row, use an empty string rather than guessing.\n"
    "- Do not include charges or notes lines as line items unless they have their own item "
    "code - but DO include lines such as delivery charges or hardener if they appear in the "
    "same numbered item list as other items.\n"
    "- Return ONLY valid JSON, no other text, no markdown code fences.\n"
    "- The JSON must have this exact shape: a top-level object with a key named line_items, "
    "whose value is an array of objects. Each object must have these exact keys: record_no, "
    "item_code, part_name, part_no, supplier, total_amount, requester, po_date. "
    "All values must be strings (wrap numbers in quotes too)."
)


# ---------- EXTRACTION HELPERS ----------

def pdf_has_text(pdf_bytes: bytes) -> bool:
    """Check if the PDF has a real, extractable text layer."""
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                if page.extract_text() and page.extract_text().strip():
                    return True
        return False
    except Exception:
        return False


def get_pdf_text(pdf_bytes: bytes) -> str:
    """Extract raw text from all pages of a text-based PDF."""
    text_parts = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            text_parts.append(f"--- Page {i + 1} ---\n{text}")
    return "\n\n".join(text_parts)


def get_pdf_page_images_b64(pdf_bytes: bytes, dpi: int = 200) -> list[str]:
    """Render each PDF page to a base64-encoded PNG image (for scanned PDFs)."""
    images = convert_from_bytes(pdf_bytes, dpi=dpi)
    b64_images = []
    for img in images:
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        b64_images.append(base64.b64encode(buf.getvalue()).decode("utf-8"))
    return b64_images


def call_claude_with_text(api_key: str, pdf_text: str) -> dict:
    client = anthropic.Anthropic(api_key=api_key)
    response = client.messages.create(
        model=MODEL,
        max_tokens=4000,
        messages=[
            {
                "role": "user",
                "content": f"{EXTRACTION_PROMPT}\n\nHere is the PO document text:\n\n{pdf_text}",
            }
        ],
    )
    return _parse_json_response(response)


def call_claude_with_images(api_key: str, b64_images: list[str]) -> dict:
    client = anthropic.Anthropic(api_key=api_key)
    content = [{"type": "text", "text": EXTRACTION_PROMPT}]
    for b64 in b64_images:
        content.append(
            {
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": "image/png",
                    "data": b64,
                },
            }
        )
    response = client.messages.create(
        model=MODEL,
        max_tokens=4000,
        messages=[{"role": "user", "content": content}],
    )
    return _parse_json_response(response)


def _parse_json_response(response) -> dict:
    text = "".join(
        block.text for block in response.content if block.type == "text"
    ).strip()
    # Strip markdown fences if present, just in case
    if text.startswith("```"):
        text = text.strip("`")
        if text.startswith("json"):
            text = text[4:]
    text = text.strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError as e:
        raise ValueError(f"Could not parse AI response as JSON: {e}\n\nRaw response:\n{text}")


def extract_po_data(api_key: str, pdf_bytes: bytes) -> tuple[list[dict], str]:
    """
    Detects whether the PDF is text-based or scanned, and routes to the
    right extraction path. Returns (line_items, method_used).
    """
    if pdf_has_text(pdf_bytes):
        pdf_text = get_pdf_text(pdf_bytes)
        result = call_claude_with_text(api_key, pdf_text)
        return result.get("line_items", []), "text"
    else:
        b64_images = get_pdf_page_images_b64(pdf_bytes)
        result = call_claude_with_images(api_key, b64_images)
        return result.get("line_items", []), "image (OCR via AI vision)"


# ---------- EXCEL EXPORT ----------

def build_excel(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "PO Line Items"

    headers = [FIELD_LABELS[f] for f in FIELDS]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for _, row in df.iterrows():
        ws.append([row.get(f, "") for f in FIELDS])

    # Auto-width columns (rough heuristic)
    for col_idx, header in enumerate(headers, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = max(
            [len(str(header))] + [len(str(v)) for v in df.get(FIELDS[col_idx - 1], [])]
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------- STREAMLIT UI ----------

st.set_page_config(page_title="PO PDF Extractor", page_icon="📄", layout="wide")

st.title("📄 PO PDF Extractor")
st.caption("Upload a Purchase Order PDF - works for both text PDFs and scanned PDFs. AI reads and extracts the fields automatically.")

# API key handling: prefer Streamlit secrets (set this up in Streamlit Cloud settings)
api_key = st.secrets.get("ANTHROPIC_API_KEY", None)

if not api_key:
    st.warning("No API key found in app secrets. Enter one below to proceed (not stored).")
    api_key = st.text_input("Anthropic API Key", type="password")

uploaded_file = st.file_uploader("Upload PO PDF", type=["pdf"])

if uploaded_file and api_key:
    pdf_bytes = uploaded_file.read()

    if "extracted_rows" not in st.session_state or st.session_state.get("last_file") != uploaded_file.name:
        with st.spinner("Reading PDF and extracting fields with AI..."):
            try:
                line_items, method = extract_po_data(api_key, pdf_bytes)
                st.session_state.extracted_rows = line_items
                st.session_state.extraction_method = method
                st.session_state.last_file = uploaded_file.name
            except Exception as e:
                st.error(f"Extraction failed: {e}")
                st.stop()

    method = st.session_state.get("extraction_method", "")
    st.success(f"Extracted {len(st.session_state.extracted_rows)} line item(s) - detected as: **{method}**")

    st.subheader("Review & edit before exporting")
    st.caption("Double-click any cell to fix mistakes (especially likely on scanned PDFs).")

    df = pd.DataFrame(st.session_state.extracted_rows)
    # Ensure all expected columns exist, in the right order
    for f in FIELDS:
        if f not in df.columns:
            df[f] = ""
    df = df[FIELDS]
    df.columns = [FIELD_LABELS[f] for f in FIELDS]

    edited_df = st.data_editor(df, num_rows="dynamic", use_container_width=True)

    # Suggest a filename based on PO number if we can find something usable,
    # otherwise fall back to a generic name.
    po_date_val = ""
    if not edited_df.empty and "PO Date" in edited_df.columns:
        po_date_val = str(edited_df["PO Date"].iloc[0]).replace("/", "-")
    default_name = f"PO_{po_date_val}.xlsx" if po_date_val else "PO_export.xlsx"

    col1, col2 = st.columns([2, 1])
    with col1:
        filename = st.text_input("Excel filename", value=default_name)
    with col2:
        st.write("")
        st.write("")
        if st.button("✅ Confirm & Prepare Download", type="primary"):
            export_df = edited_df.copy()
            export_df.columns = FIELDS  # map back to internal keys for build_excel
            excel_bytes = build_excel(export_df)
            st.session_state.excel_ready = excel_bytes
            st.session_state.excel_filename = filename if filename.endswith(".xlsx") else f"{filename}.xlsx"

    if st.session_state.get("excel_ready"):
        st.download_button(
            label="⬇️ Download Excel",
            data=st.session_state.excel_ready,
            file_name=st.session_state.excel_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

elif uploaded_file and not api_key:
    st.info("Enter your Anthropic API key above to process this file.")
else:
    st.info("Upload a PO PDF to get started.")
